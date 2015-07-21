from openpyxl import Workbook, load_workbook
import os

try:
    summary = load_workbook('_FGT Summary.xlsx')
except IOError as e:
    summary = Workbook()
    summary.active.title = "FGT Summary"
    ws = summary.active
    ws.cell('A1').value = "Filename"
    ws.cell('B1').value = "Service Type"
    ws.cell('C1').value = "# of ppl"
    ws.cell('D1').value = "MPP Cont. %"
    ws.cell('E1').value = "MPP Part. %"
    ws.cell('F1').value = "LTD %"
    ws.cell('G1').value = "FTE"
    ws.cell('H1').value = "Total Cost of Contract"
    ws.cell('I1').value = "Amt BACI receives"
    ws.cell('J1').value = "Gap"
    ws.cell('K1').value = "Input Value of Amt Received"
    ws.cell('L1').value = "Proposed Funding Changes"

path = 'excelfiles'
contracts = []

for filename in os.listdir(path): #go through each contract file in directory
    wb = load_workbook(path + "/" + filename, data_only=True) #open a contract file
    
    if 'HS' not in filename: #Not Home Share Contracts
   	row = []
    	sheet = wb.get_sheet_by_name('Page1') 
        row.append(filename)
    	row.append(sheet.cell('J13').value) # servicetype
    	row.append(sheet.cell('D15').value) # num of ppl
    	row.append(sheet.cell('D45').value) # MPP Cont %
    	row.append(sheet.cell('D47').value) # MPP Part %
    	row.append(sheet.cell('D49').value) # LTD %

    	sheet = wb.get_sheet_by_name('Page2A')
    	row.append(sheet.cell('U8').value) #FTE

    	sheet = wb.get_sheet_by_name('Page3')
    	row.append(sheet.cell('E90').value) # Total Cost of Contract
    	row.append(sheet.cell('E107').value) # Amt BACI receives
    	row.append(sheet.cell('E98').value) # Gap
    	row.append(sheet.cell('E109').value) # Input Value of Amt Received
    	row.append(sheet.cell('E110').value) # Proposed Funding Changes"""
        contracts.append(row)

    else: #Home Share Contracts
    	row = []

    	sheet = wb.get_sheet_by_name('Page 1')
        row.append(filename)
    	row.append(sheet.cell('I9').value) # servicetype
    	row.append(sheet.cell('C15').value) # number of ppl
    	row.append(sheet.cell('C29').value) # MPP Cont %
    	row.append(sheet.cell('C31').value) # MPP Part %
    	row.append(sheet.cell('C35').value) # LTD %

    	sheet = wb.get_sheet_by_name('Page 5')
    	row.append(sheet.cell('G10').value) # FTE

    	sheet = wb.get_sheet_by_name('Page 2')
    	row.append(sheet.cell('E39').value) # Total Cost of Contract
    	row.append(sheet.cell('E43').value) # Amt BACI receives
    	row.append('see proposed changes') # Gap
    	row.append(sheet.cell('E43').value) # Input Value of Amt Received
    	row.append(sheet.cell('E44').value) # Proposed Funding Changes
    	contracts.append(row)
print contracts

#writing the lists to the excel file
x = -1
y = -1
for i in range(2, 2+len(contracts)):
    x = x + 1
    for j in range(1,1+len(contracts[0])):
        y = y + 1
        c = summary.active.cell(row = i, column = j)
        c.value = contracts[x][y]
    y= -1
    
summary.save(filename = '_FGT Summary.xlsx') #saving the excel summary file as '_FGT Summary.xlsx'
